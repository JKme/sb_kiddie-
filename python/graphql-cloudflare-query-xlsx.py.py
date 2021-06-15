# coding: utf-8
import requests
import json
import openpyxl
from pathlib import Path
import csv

EMAIL = ""
AUTH = ""
URL = "https://api.cloudflare.com/client/v4/graphql/"
ZONE = ""
N = "/"

Headers = {
    "X-Auth-Email": EMAIL,
    "X-Auth-Key": AUTH,
    "Content-Type": "application/json"
}


def query_payload(ip):
    d = f'''{{
  viewer {{
    zones(filter: {{zoneTag: "{ZONE}"}}) {{
      firewallEventsAdaptiveGroups(filter: {{datetime_geq: "2021-06-06T12:00:00Z", datetime_leq: "2021-06-09T09:55:00Z",clientIP:"{ip}", action_in: ["connection_close","block"]}}, limit: 5000) {{ 
        count
        dimensions {{
        action
        clientRequestHTTPHost
        clientAsn
        clientCountryName
        clientIP
        clientRequestPath
        clientRequestQuery
        datetime
        rayName
        source
        userAgent
            }}
          }}
        }}
      }}
    }}'''
    return d


def write(file, dst_file):
    xlsx_file = Path(file)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    with open(dst_file, 'w') as f:
        c = csv.writer(f)
        c.writerow(["ZID", "IP", "是否主动自动拦截", "首次主动拦截时间", "首次主动拦截原因", "最后主动拦截时间","最后主动拦截原因", "拦截原因", "IP是否在黑名单"])  # 写入第一行
        for row in sheet.iter_rows(2, sheet.max_row):
            row_0 = row[0].value
            ip = row[1].value
            print(f"[*]: Checking: {ip}")
            l = waf_query(ip)
            r = [row_0, ip]
            r.extend(l)
            # print(r)
            c.writerow(r)


def waf_query(ip):
    """
    查询IP是否在WAF拦截日志里面
    :param ip:
    :return:
    """
    block_flag = block_time_start = block_time_end = block_source1 = block_source2 = ip_black ="/"
    payload = query_payload(ip)
    ret = requests.post(URL, headers=Headers, data=json.dumps({"query": payload})).content
    rs = json.loads(ret)["data"]["viewer"]["zones"][0]["firewallEventsAdaptiveGroups"]
    if rs:
        print(rs)
        block_time_start = rs[0]["dimensions"]["datetime"]
        block_source1 = rs[0]["dimensions"]["source"]
        block_time_end = rs[-1]["dimensions"]["datetime"]
        block_source2 = rs[-1]["dimensions"]["source"]
        block_flag = "是"
    r1 = ip_rule_query(ip)
    if r1:
        ip_black = "是"
    l = [block_flag, block_time_start, block_source1, block_time_end, block_source2, ip_black]
    print(l)
    return l


def ip_rule_query(ip):
    """
    查询IP是否在黑名单，黑名单返回true，白false
    :param ip:
    :return:
    """
    ret = False
    email = ""
    auth = ""
    headers = {
        "X-Auth-Email": email,
        "X-Auth-Key": auth,
        "Content-Type": "application/json"
    }
    url = f"https://api.cloudflare.com/client/v4/accounts/{zone}/firewall/access_rules/rules"
    data = {"page": 1, "per_page": 10, "match": "any", "configuration_value": f"contains:{ip}",
            "notes": f"contains:{ip}"}
    r = requests.get(url, params=data, headers=headers)
    if json.loads(r.text)["result"]:
        ret = True
    return ret


if __name__ == "__main__":
    write("XID.xlsx", "cloudflare_block_record.csv")
